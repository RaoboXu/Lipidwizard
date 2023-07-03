import os.path
from typing import List
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet


def CreateEmptyXlsx(name:str):
    book = xl.Workbook()
    book.save(name)
    

def checkTableTitles(tableTitles: List, targetTitles: List):

    for col in targetTitles:
        if not tableTitles.count(col):
            return False
        return True
    # if len(tableTitles) == len(targetTitles):
    #     for i in range(len(tableTitles)):
    #         if str(tableTitles[i]).isascii() and str(targetTitles[i]).isascii():
    #             if tableTitles[i].lower() != targetTitles[i].lower():
    #                 return False
    #         else:
    #             if tableTitles[i] != targetTitles[i]:
    #                 return False
    #     return True
    # else:
    #     return False

def getSheetNames(source: str):
    if os.path.exists(source):
        workbook = xl.load_workbook(source, read_only=True, data_only=True)
        names = workbook.get_sheet_names()
        return names
    return []

def readXlsx(source: str, data_result: list, sheet_name="", headers: list[str] = list()):
    if os.path.exists(source):
        workbook = xl.load_workbook(source, read_only=True, data_only=True)
        # print("file loaded")
        sheet:Worksheet=None
        if sheet_name == "":
            sheet = workbook.active
            print('Reading:\t'+source)
        else:
            sheet = workbook.get_sheet_by_name(name=sheet_name)
            print('Reading:\t{}/{}'.format(source,sheet_name))
        # Read all data
        if len(headers) > 0:
            row0 = sheet[1]
            titles = list()
            for cell in row0:
                titles.append(str(cell.value))
            print(titles)
            if(checkTableTitles(titles, headers) is False):
                print(source + 'has no content you want to read')
                return "FAILED"
        i = 0
        for row in sheet.rows:
            data_row = []
            for cell in row:
                data_row.append(cell.value)
            data_result.append(data_row)
            i=(i+1)%1e6
            if(i>10000):
                print('\Reading'+'.'*int(i/10000),end=' ')
        workbook.close()
        print("Reading Success!")
        return 'SUCCESS'
    else:
        print('file: \''+str(source)+'\' does not exist')


# def readXlsx(source: str, data_result: list, sheet_name="", headers: list[str] = list()):
#     if os.path.exists(source):
#         workbook = xl.load_workbook(source, read_only=False, data_only=True)
#         # print("file loaded")
#         sheet:Worksheet=None
#         if sheet_name == "":
#             sheet = workbook.active
#         else:
#             sheet = workbook.get_sheet_by_name(name=sheet_name)
        
#         # Read all data
#         print('Reading:\t'+source)
#         if len(headers) > 0:
#             row0 = sheet[1]
#             titles = list()
#             for cell in row0:
#                 titles.append(str(cell.value))
#             if(checkTableTitles(titles, headers) is False):
#                 print(source + 'has no content you want to read')
#                 return "FAILED"
#         i = 0.
#         first_col = sheet['A']
#         n = len(first_col)
    
#         for row in sheet.rows:
#             data_row = []
#             for cell in row:
#                 data_row.append(cell.value)
#             data_result.append(data_row)
#             i += 1
#             print('\r{:.2f}%'.format((i/n)*100), end=' ')
#         workbook.close()
#         print('\n'+source + ' loaded ')
#         return 'SUCCESS'
#     else:
#         print('file: \''+str(source)+'\' does not exist')


def SaveDataXlsx(target: str, data: List, titles: List[str] = list()):
    print('Writing data to file:'+target)
    # Write Data
    outbook = xl.Workbook(write_only=True)
    outsheet = outbook.create_sheet()
    # Write title
    if(len(titles) > 0):
        outsheet.append(titles)
    for data_row in data:
        outsheet.append(data_row)
    outbook.save(target)
    outbook.close()
    print('Writing data'+target+' SUCCESS!')

def SaveDataXlsxWorkSheet(workbook: str, worksheet:str, data: List, titles: List[str] = list()):
    print('Writing '+worksheet)
    # Write Data
    # outbook = xl.Workbook(write_only=True)
    # outsheet = outbook.create_sheet()
    if not os.path.exists(workbook):
        print('workbook {} not exists'.format(workbook))
        return 'FAILED'
    outbook = xl.load_workbook(workbook)
    outsheet = outbook.create_sheet(worksheet)

    # Write title
    if(len(titles) > 0):
        outsheet.append(titles)
    for data_row in data:
        outsheet.append(data_row)
    outbook.save(workbook)
    outbook.close()
    print('Writing data'+worksheet+' SUCCESS!')
    return 'SUCCESS'




if __name__ == '__main__':
    print('test')
    # data = [[1,2,3,4,5,6,7],[11,22,33,44,55,66,77]]
    # SaveDataXlsx('./test/test.xlsx',data=data)
    # SaveDataXlsxWorkSheet('./test/test.xlsx','sheet2',data=data)
    CreateEmptyXlsx('./test/test_1.xlsx')
